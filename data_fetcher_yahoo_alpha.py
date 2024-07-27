"""
Dieses Skript lädt die Kursdaten der bedeutensten Aktienindizes von China (Shanghai Composite Index), Großbritannien(FTSE 100 Index), Frankreich(CAC 40 Index), 
Russland(MOEX Russia Index), Brasilien(Bovespa Index - Ibovespa), Japan(Nikkei 225), Deutschland(Dax) und Amerika(S&P500) herunter.
Danach werden sie in eine richtige Form gebracht, bereinigt und interpoliert. Im Anschluss werden Sie in mehrere Excell files geschrieben
 (kann auch zu csv geändert werden). Außerdem kann es auch zum updaten der Daten verwendet werden.

Die Daten werden im Format:
date open high low close
abgespeichert
"""

import requests
import csv
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

#mit alphavantage sind 5 API Requests in der Minute erlaubt und 100 am Tag

API_KEY = 'f772cb1126msh67c279955f8b42bp10c102jsn5c6f540f61ce'
FILE_PATH = r'C:\Users\arne\Desktop\Informatik\Data Science Projekt\Data\indize_data'
HEADER = ['open', 'high', 'low', 'close', 'volume']

dict_yahoo = {'Shenzhen Index' : '399001.SZ', 'FTSE 100 Index' : '^FTSE', 'MOEX Russia Index' : 'IMOEX.ME', 'Bovespa Index' : '^BVSP',
                 'Nikkei 225' : '^N225', 'S&P500' : '^GSPC', 'TSX Composite Index' : '^GSPTSE', 'NSE Nifty 50' : '^NSEI', 
                 'MERVAL Index' : '^MERV', 'ASX 200' : '^AXJO', 'IDX Composite Index' : '^JKSE', 'BMV IPC' : '^MXX', 
                 'KOSPI Index' : '^KS11'}


#url = 'https://www.alphavantage.co/query?function=TIME_SERIES_DAILY&symbol=GOLD&apikey=f772cb1126msh67c279955f8b42bp10c102jsn5c6f540f61ce&outputsize=full'
#Diese URL ist für alle Aktien und ETFS
url_stocks = 'https://www.alphavantage.co/query?function=TIME_SERIES_DAILY&symbol='

dict_stocks = {'DAX' : 'DAX', 'CAC 40 Index' : 'CAC'}



#Es passiert gelegentlich, dass Nullwerte von der API kommen, diese müssen dann aufgefüllt werden
def nullwerteAuffuellen(all_stock_data):
    print('Nullwerte werden mit den vorherigen Werten gefüllt')
    for symbol_key in all_stock_data.keys():
        date_list = list(all_stock_data[symbol_key].keys())
        for i in range(len(date_list)):
            row = all_stock_data[symbol_key][date_list[i]]

            if i != 0:
                old_row = all_stock_data[symbol_key][date_list[i - 1]]
            else:
                old_row = row

            if row['1. open'] == 0:
                all_stock_data[symbol_key][date_list[i]] = old_row


def cutDates(all_stock_data):
    print('Suche nach dem jüngstem Datum gestartet')
    # Es wird nach dem jüngstem Datum gesucht
    newest_date = '1900-10-25'
    for symbol_key in all_stock_data.keys():
        stock = all_stock_data[symbol_key]
        date = list(stock.keys())[len(list(stock.keys())) - 1]

        if datetime.strptime(newest_date, '%Y-%m-%d') < datetime.strptime(date, '%Y-%m-%d'):
            newest_date = date

    # alles wegcuten, was älter als newest_date ist
    stock_keys = list(all_stock_data.keys())[:]

    for symbol_key in stock_keys:
        stock = all_stock_data[symbol_key]
        del all_stock_data[symbol_key]

        # aufpassen, dass newest_date in jedem dict vorhanden ist
        new_stock = {}
        for date in stock.keys():
            if datetime.strptime(newest_date, '%Y-%m-%d') <= datetime.strptime(date, '%Y-%m-%d'):
                new_stock[date] = stock[date]

        all_stock_data[symbol_key] = new_stock

#Diese Funktion holt sich alle verfügbaren historischen Daten und schreibt sie in ein File
def getAlLStockDataAndWriteItToFile():
    all_stock_data = getAllData('full')

    cutDates(all_stock_data)

    nullwerteAuffuellen(all_stock_data)

    print('Werte werden interpoliert')

    print('Referenzobjekt erstellen')

    # Hierdurch wird dir performanz deutlch erhöht
    first_element_dict = {}
    first_element_dict[list(all_stock_data.keys())[0]] = all_stock_data[list(all_stock_data.keys())[0]]

    referenz = interpolateValues(all_stock_data, first_element_dict)

    all_stock_data = interpolateValues(referenz, all_stock_data)

    # Daten in XLSX File schreiben
    for symbol_key in all_stock_data.keys():
        writeXLSXFiles(symbol_key, convertDict(all_stock_data[symbol_key]))

    print("Daten wurden erfolgreich angelegt.")

#Diese Funktion updatet schon vorhandene Daten Tabellen
def updateData():
    print("Starte mit dem updaten der Daten")

    all_stock_data = getAllData('compact')

    cutDates(all_stock_data)

    nullwerteAuffuellen(all_stock_data)

    print('Werte werden interpoliert')

    print('Referenzobjekt erstellen')

    # Hierdurch wird dir performanz deutlch erhöht
    first_element_dict = {}
    first_element_dict[list(all_stock_data.keys())[0]] = all_stock_data[list(all_stock_data.keys())[0]]

    referenz = interpolateValues(all_stock_data, first_element_dict)

    all_stock_data = interpolateValues(referenz, all_stock_data)

    all_new_data = {}
    for symbol_key in all_stock_data.keys():
       old_data = readXLSXFile(FILE_PATH + '\\' + symbol_key + '.xlsx')
       new_data = all_stock_data[symbol_key]
       all_new_data[symbol_key] = new_data | old_data

    # Daten in XLSX File schreiben
    for symbol_key in all_new_data.keys():
        writeXLSXFiles(symbol_key, convertDict(all_new_data[symbol_key]))




#liest die Daten von einer XLSX Datei ein und konvertiert sie direkt in das richtige Format
def readXLSXFile(path):
    wb = load_workbook(path)
    sheet = wb.active
    row_count = sheet.max_row

    data = {}
    for i in range(1, row_count + 1):
        data[sheet.cell(row=i, column=1).value] = {'1. open': sheet.cell(row=i, column=2).value,
                                                   '2. high': sheet.cell(row=i, column=3).value,
                                                   '3. low': sheet.cell(row=i, column=4).value,
                                                   '4. close': sheet.cell(row=i, column=5).value}

    return data


#schreibt die Daten in ein XLSX File
def writeXLSXFiles(symbol, data):
    workbook = Workbook()
    sheet = workbook.active

    for row in range(len(data)):
        sheet['A'+str(row+1)] = data[row][0]
        sheet['B' + str(row + 1)] = data[row][1]
        sheet['C' + str(row + 1)] = data[row][2]
        sheet['D' + str(row + 1)] = data[row][3]
        sheet['E' + str(row + 1)] = data[row][4]

    path = FILE_PATH + '\\' + symbol + '.xlsx'
    workbook.save(path)




#schreibt die daten in ein CSV File
def writeCSVFiles(symbol, data):
    path = FILE_PATH + '\\' + symbol + '.csv'
    file = open(path, 'w+', newline='')

    writer = csv.writer(file)

    for row in data:
        writer.writerow(row)
    #writer.writerows(data)

#macht aus einem dict eine liste, um die dann in ein CSV File zu schreiben
def convertDict(dict):
    list = []

    for key_date in dict.keys():
        row_entry = [key_date]
        for key_value in dict[key_date].keys():
            row_entry.append(dict[key_date][key_value])
        list.append(row_entry)

    return list

#fügt alle historischen Aktiendaten in einem dictonary zusammen
def getAllData(outputsize):
    bigdata = {}

    #Get Data for Stocks
    for key in dict_stocks.keys():
        json = sendRequestAlpha(dict_stocks[key], url_stocks, outputsize)

        #format data
        del json['Meta Data']
        new_data = {}

        for date in json['Time Series (Daily)'].keys():
            new_data[date] = json['Time Series (Daily)'][date]
            del new_data[date]['5. volume']

        bigdata[key] = new_data
        print('Daten von: ' + key + ' erhalten')


    #Get Data for Features from yahoo
    for key in dict_yahoo.keys():
        json = sendRequestYahoo(dict_yahoo[key])

        del json['meta']

        new_dict = {}
        for utc_time in dict(reversed(json['items'].items())).keys():
            temp = json['items'][utc_time]
            new_dict[convertDate(temp['date'])] = {'1. open' : temp['open'], '2. high' : temp['high'], '3. low' : temp['low'], '4. close' : temp['close']}

        bigdata[key] = new_dict
        print('Daten von: ' + key + ' erhalten')

    return bigdata

#schickt request an die yahoo API und gibt die Antwort als json zurück
def sendRequestYahoo(symbol):
    url = "https://yahoo-finance15.p.rapidapi.com/api/yahoo/hi/history/" + symbol + "/1d"

    querystring = {"diffandsplits": "false"}

    headers = {
        "X-RapidAPI-Key": "f772cb1126msh67c279955f8b42bp10c102jsn5c6f540f61ce",
        "X-RapidAPI-Host": "yahoo-finance15.p.rapidapi.com"
    }

    response = requests.request("GET", url, headers=headers, params=querystring)

    return response.json()

#schickt request an die alpha vantage API und gibt die Antwort als json zurück
def sendRequestAlpha(symbol, url, outputsize):
    url = url + symbol + '&apikey=f772cb1126msh67c279955f8b42bp10c102jsn5c6f540f61ce&outputsize=' + outputsize
    request = requests.get(url)
    time.sleep(12) # Wird benötigt um nicht mehr als 5 API Requests in der Minute loszuschicken
    return request.json()

#this function takes a date from the yahoo API and converts in a alpha date
def convertDate(date):
    return date[6:] + "-" + date[:2] + "-" + date[3:5]

def interpolateValues(compareDataSet, interpoleDataSet):
    # Werte interpolieren (fehlende Werte mit Durchschnittswerten auffüllen)
    for symbol_key in compareDataSet.keys():

        stock_data = compareDataSet[symbol_key]

        for date_key_to_check in stock_data.keys():

            for symbol_key_next in interpoleDataSet.keys():
                if symbol_key_next != symbol_key:

                    list_keys = list(interpoleDataSet[symbol_key_next].keys())
                    # gucken, ob date_key_to_check vorhanden ist
                    if date_key_to_check not in list_keys:
                        # gucken, wo der fehlende wert eingetragen werden muss
                        for i in range(len(list_keys)-1):
                            if i+1 == len(list_keys):
                                print(list_keys[i+1])
                            if datetime.strptime(list_keys[i], '%Y-%m-%d') > datetime.strptime(date_key_to_check,
                                                                                               '%Y-%m-%d') > datetime.strptime(
                                    list_keys[i + 1], '%Y-%m-%d'):
                                row_1 = interpoleDataSet[symbol_key_next][list_keys[i]]
                                row_2 = interpoleDataSet[symbol_key_next][list_keys[i + 1]]
                                new_entry = {}
                                for row_key in row_1.keys():
                                    new_entry[row_key] = (float(row_1[row_key]) + float(row_2[row_key])) / 2
                                # neuen eintrag an richtiger stelle einfügen, indem ein neues dict verwendet wird
                                new_dict = {}
                                for k in range(len(list_keys)):
                                    if k == i + 1:
                                        new_dict[date_key_to_check] = new_entry
                                        new_dict[list_keys[k]] = interpoleDataSet[symbol_key_next][list_keys[k]]
                                    else:
                                        new_dict[list_keys[k]] = interpoleDataSet[symbol_key_next][list_keys[k]]
                                interpoleDataSet[symbol_key_next] = new_dict
                                # 1, open, 2, high, 3, low, 4, close

    return interpoleDataSet

#updateData()
getAlLStockDataAndWriteItToFile()